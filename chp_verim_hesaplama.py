import PySimpleGUI as sg
from openpyxl import Workbook, load_workbook

sg.theme("Dark")

# GİRİŞ İÇİN CELL VE EXCEL FİLE BELİRLE
cell_elc = ["B10","C10","D10","E10","F10","G10","H10","I10","J10","K10","L10","M10"]
cell_ngas = ["B11","C11","D11","E11","F11","G11","H11","I11","J11","K11","L11","M11"]
working_cell = ["L4","L5","L6"]
unit_cell = ["I3","I4","I5"]
default_value = {'January Electric': '127930', 'January Natural gas': '17564', 'February Electric': '102663', 'Februar Natural gas': '15502', 'March Electric': '131908', 'March Natural gas': '14590', 'April Electric': '116119', 'April Natural gas': '13561', 'May Electric': '123626', 'May Natural gas': '8504', 'June Electric': '143693', 'June Natural gas': '5156', 'July Electric': '138568', 'July Natural gas': '4489', 'August Electric': '154250', 'August Natural gas': '4980', 'September Electric': '145153', 'September Natural gas': '6567', 'October Electric': '134999', 'October Natural gas': '6990', 'November Electric': '132985', 'November Natural gas': '9314', 'December Electric': '104900', 'December Natural gas': '9608', 'T1 Work Time': '11', 'T2 Work Time': '5', 'T3 Work Time': '8', '-FolderDirectory-': '', 'Browse': ''}
excel_file = "assets/template.xlsx"
output_excel_file = "CHAP_System_Table.xlsx"
default_data = 0

wb = load_workbook(excel_file)
ws = wb.active


layout = [

    [sg.Text("Please fill in the blanks with the \nmonthly electricity(kWh) and Natural Gas(m3) consumption")],
    [sg.Text("", size=(15,1)), sg.Text("Electricity", size=(15,1), justification='center'), sg.Text("Natural Gas", size=(15,1))],
    [sg.Text("January:", size=(15,1)), sg.InputText(key = "January Electric" , size=(15,1)), sg.InputText(key = "January Natural gas", size=(15,1))],
    [sg.Text("February:", size=(15,1)), sg.InputText(key = "February Electric" , size=(15,1)), sg.InputText(key = "Februar Natural gas", size=(15,1))],
    [sg.Text("March:", size=(15,1)), sg.InputText(key = "March Electric", size=(15,1)), sg.InputText(key = "March Natural gas", size=(15,1))],
    [sg.Text("April:", size=(15,1)), sg.InputText(key = "April Electric", size=(15,1)), sg.InputText(key = "April Natural gas", size=(15,1))],
    [sg.Text("May:", size=(15,1)), sg.InputText(key = "May Electric", size=(15,1)), sg.InputText(key = "May Natural gas", size=(15,1))],
    [sg.Text("June:", size=(15,1)), sg.InputText(key = "June Electric", size=(15,1)), sg.InputText(key = "June Natural gas", size=(15,1))],
    [sg.Text("July:", size=(15,1)), sg.InputText(key = "July Electric", size=(15,1)), sg.InputText(key = "July Natural gas", size=(15,1))],
    [sg.Text("August:", size=(15,1)), sg.InputText(key = "August Electric", size=(15,1)), sg.InputText(key = "August Natural gas", size=(15,1))],
    [sg.Text("September:", size=(15,1)), sg.InputText(key = "September Electric", size=(15,1)), sg.InputText(key = "September Natural gas", size=(15,1))],
    [sg.Text("October:", size=(15,1)), sg.InputText(key = "October Electric", size=(15,1)), sg.InputText(key = "October Natural gas", size=(15,1))],
    [sg.Text("November:", size=(15,1)), sg.InputText(key = "November Electric", size=(15,1)), sg.InputText(key = "November Natural gas", size=(15,1))],
    [sg.Text("December:", size=(15,1)), sg.InputText(key = "December Electric", size=(15,1)), sg.InputText(key = "December Natural gas", size=(15,1))],
    
    [
    sg.Text("Work Hour Daily:", size=(15,1)),
    sg.Combo(['0','1','2', '3','4','5','6','7','8','9','10','11'], default_value='06-17',key='T1 Work Time', size=(8,1)),
    sg.Combo(['0','1','2', '3','4','5'], default_value='17-22',key='T2 Work Time', size=(7,1)),
    sg.Combo(['0','1','2', '3','4','5','6','7','8'], default_value='22-06',key='T3 Work Time', size=(8,1)),
    ],

    [sg.Text("Output Directory:", size=(15,1)), sg.Input(key = "-FolderDirectory-", disabled= True, text_color="black", size=(32,1)), sg.FolderBrowse(button_text= "Browse", key= "Browse")],

    [sg.Submit(), sg.Button("Clear"),sg.Button("Default") ,sg.Exit(), sg.Push(), sg.Button("Credits")]
    
        ]

window = sg.Window("Combined Heat And Power System Calculator", layout)

def clear_input():

    for key in value:
        if key != "Browse":
            window[key].update("")
    return None

def value_check(dict):
    
    key = list(dict.keys())
    usage = list(dict.values())
    wrong_months = []

    i = 0
    for x in usage:
        try:
            y = float(x)
            print(f"{y} passed", end=", ")
            i = i + 1
        except ValueError:
            wrong_months.append(key[i])
            print(f"{x} failed", end=", ")
            i = i + 1
    print("\n------------------------------")
    return wrong_months

def co_unit_finder(electric):
    if electric <= 100:
        co_gen_unit = [70, 109, 204]
    elif electric <= 200:
         co_gen_unit = [155, 186, 377]
    elif electric <= 300:
         co_gen_unit = [206, 246, 495]
    elif electric <= 400:
         co_gen_unit = [331, 392, 789]
    elif electric <= 500:
         co_gen_unit = [430, 640, 1160]
    elif electric <= 600:
         co_gen_unit = [528, 705, 1344]
    elif electric <= 700:
         co_gen_unit = [600, 717, 1418]
    elif electric <= 900:
         co_gen_unit = [800	, 952, 1882]
    elif electric <= 1200:
         co_gen_unit = [1200, 1428, 2818]
    elif electric <= 1800:
         co_gen_unit = [1560, 1884, 3696]
    elif electric <= 2200:
         co_gen_unit = [2000, 2,372, 4690]
    elif electric <= 4000:
         co_gen_unit = [3333, 3740, 7650]
    elif electric <= 7000:
         co_gen_unit = [4500, 4904, 10160]
    elif electric >= 7000:
         co_gen_unit = [10426, 9825, 22176]

    return co_gen_unit


while True:
    event, value = window.read()

    if event == sg.WIN_CLOSED or event == "Exit":
        break
    
    if event == "Clear":
        clear_input()
    
    if event == "Default":
        default_data = 1
        sg.popup("Default Values set. Please make sure to enter Output Directory and press Submit button to use default values.")
    
    if event == "Credits":
        sg.popup("Desinged by \nKaan, Esat, Emre\nDoç.Dr. Mustafa BAYSAL ")

    if event == "Submit":
        
        output_folder = value["-FolderDirectory-"]
        
        if default_data == 1:
            value = default_value
            value["-FolderDirectory-"] = output_folder

        work_hour_list= [value["T1 Work Time"], value["T2 Work Time"], value["T3 Work Time"],]

        value.pop("Browse")
        value.pop("-FolderDirectory-")
        

        if value_check(value) == [] and output_folder != "":
            value.pop("T1 Work Time")
            value.pop("T2 Work Time")
            value.pop("T3 Work Time")

            save_spot = output_folder + "/" + output_excel_file
            
            k = 0
            for x in work_hour_list:
                ws[working_cell[k]].value = (int(x) * 30) - 10
                wb.save(save_spot)
                k = k + 1

            all_usage = list(value.values())
            i = 0
            j = 0

            ort_elektrik = 0
            for x in all_usage:
                if i % 2 == 0:
                    ws[cell_elc[j]].value = int(x)
                    wb.save(save_spot)
                    ort_elektrik = ort_elektrik + (int(x)/ ((int(work_hour_list[0])*1 + int(work_hour_list[1])*85/100 + int(work_hour_list[2])*65/100)*30))
                else:
                    ws[cell_ngas[j]].value = int(x)
                    wb.save(save_spot)
                    j = j + 1
                i = i + 1


            if default_data == 0:
                ort_elektrik = ort_elektrik / 12
                co_unit_list = co_unit_finder(ort_elektrik)
            elif default_data == 1:
                co_unit_list = [206, 246, 495]

            i = 0
            for x in co_unit_list:
                ws[unit_cell[i]].value = int(x)
                wb.save(save_spot)
                i = i + 1
            
            sg.popup("Document Saved to Output Directory ")
            window.close()

        else:
            
            wrong_entry = value_check(value)
            
            if output_folder == "":
                wrong_entry.append("Output Directory")

            sg.popup(f"The following sections have been filled in incorrectly or incompletely, please try again.\n\n--> {', '.join(wrong_entry)}")

            
